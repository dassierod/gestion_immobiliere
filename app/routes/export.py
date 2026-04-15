from flask import Blueprint, send_file, request
from flask_login import login_required
from ..models import Room, Payment, Expense, Parameter, MONTHS
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

export_bp = Blueprint("export", __name__, url_prefix="/export")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHEADER_FONT = Font(bold=True)
PAID_FILL = PatternFill("solid", fgColor="C6EFCE")
LATE_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply(cell, fill=None, font=None, align=CENTER, border=BORDER):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = align
    cell.border = border


@export_bp.route("/excel")
@login_required
def excel():
    year = request.args.get("year", datetime.now().year, type=int)
    rooms = Room.query.filter_by(is_active=True).order_by(Room.id).all()

    caissier_param = Parameter.query.filter_by(key="caissier_mensuel").first()
    caissier_default = caissier_param.as_float() if caissier_param else 0.0

    wb = openpyxl.Workbook()

    # --- Paramètres sheet ---
    ws_params = wb.active
    ws_params.title = "Paramètres"
    params = Parameter.query.order_by(Parameter.id).all()
    ws_params.append(["Clé", "Valeur", "Description"])
    for cell in ws_params[1]:
        _apply(cell, fill=HEADER_FILL, font=HEADER_FONT)
    for i, p in enumerate(params, start=2):
        ws_params.append([p.key, p.value, p.label])
    for col in ws_params.columns:
        ws_params.column_dimensions[get_column_letter(col[0].column)].width = 28

    # --- Suivi sheet ---
    ws = wb.create_sheet(title=f"Suivi {year}")

    # Build header rows
    # Row 1: group headers
    # Row 2: column sub-headers
    col = 1
    ws.cell(row=1, column=col, value="Mois")
    ws.cell(row=2, column=col, value="Mois")
    _apply(ws.cell(row=1, column=col), fill=HEADER_FILL, font=HEADER_FONT)
    _apply(ws.cell(row=2, column=col), fill=HEADER_FILL, font=HEADER_FONT)
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    ws.column_dimensions[get_column_letter(col)].width = 14
    col += 1

    room_col_start = {}
    for room in rooms:
        room_col_start[room.id] = col
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        header_cell = ws.cell(row=1, column=col, value=room.name)
        _apply(header_cell, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        for sub, label in enumerate(["Statut", "Attendu", "Payé", "Solde dû"]):
            c = ws.cell(row=2, column=col + sub, value=label)
            _apply(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
            ws.column_dimensions[get_column_letter(col + sub)].width = 14
        col += 4

    # Totals group
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    _apply(ws.cell(row=1, column=col, value="Totaux"), fill=HEADER_FILL, font=HEADER_FONT)
    for sub, label in enumerate(["Total Attendu", "Total Payé", "Total Solde dû"]):
        c = ws.cell(row=2, column=col + sub, value=label)
        _apply(c, fill=HEADER_FILL, font=HEADER_FONT)
        ws.column_dimensions[get_column_letter(col + sub)].width = 16
    totals_col = col
    col += 3

    # Dépenses group
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    _apply(ws.cell(row=1, column=col, value="Dépenses"), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    for sub, label in enumerate(["Hygiène", "Caissier"]):
        c = ws.cell(row=2, column=col + sub, value=label)
        _apply(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
        ws.column_dimensions[get_column_letter(col + sub)].width = 14
    hygiene_col = col
    col += 2

    ws.cell(row=1, column=col, value="Total Dépenses")
    ws.cell(row=2, column=col, value="Total Dépenses")
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    _apply(ws.cell(row=1, column=col), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    ws.column_dimensions[get_column_letter(col)].width = 16
    total_dep_col = col
    col += 1

    # Dépôt group
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    _apply(ws.cell(row=1, column=col, value="Dépôt"), fill=HEADER_FILL, font=HEADER_FONT)
    for sub, label in enumerate(["Dépôt conseillé", "Dépôt réel", "Écart"]):
        c = ws.cell(row=2, column=col + sub, value=label)
        _apply(c, fill=HEADER_FILL, font=HEADER_FONT)
        ws.column_dimensions[get_column_letter(col + sub)].width = 18
    depot_col = col
    col += 3

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # Data rows
    for month_num in range(1, 13):
        row = month_num + 2
        payments = {p.room_id: p for p in Payment.query.filter_by(year=year, month=month_num).all()}
        expense = Expense.query.filter_by(year=year, month=month_num).first()

        ws.cell(row=row, column=1, value=MONTHS[month_num - 1]).border = BORDER

        total_attendu = 0.0
        total_paye = 0.0
        for room in rooms:
            c0 = room_col_start[room.id]
            p = payments.get(room.id)
            attendu = room.monthly_rent
            paye = p.amount_paid if p else 0.0
            solde = attendu - paye
            status = (p.status if p else ("En retard" if attendu > 0 else "—"))
            total_attendu += attendu
            total_paye += paye

            status_cell = ws.cell(row=row, column=c0, value=status)
            status_fill = PAID_FILL if status == "Payé" else (LATE_FILL if status == "En retard" else PARTIAL_FILL)
            _apply(status_cell, fill=status_fill)
            for sub, val in enumerate([attendu, paye, solde], start=1):
                _apply(ws.cell(row=row, column=c0 + sub, value=val))

        total_solde = total_attendu - total_paye
        for sub, val in enumerate([total_attendu, total_paye, total_solde]):
            _apply(ws.cell(row=row, column=totals_col + sub, value=val))

        hygiene = expense.hygiene if expense else 0.0
        caissier = expense.caissier if expense else caissier_default
        total_depenses = hygiene + caissier
        depot_conseille = total_paye - total_depenses
        depot_reel = expense.depot_reel if expense else 0.0
        ecart = depot_reel - depot_conseille

        _apply(ws.cell(row=row, column=hygiene_col, value=hygiene))
        _apply(ws.cell(row=row, column=hygiene_col + 1, value=caissier))
        _apply(ws.cell(row=row, column=total_dep_col, value=total_depenses))
        _apply(ws.cell(row=row, column=depot_col, value=depot_conseille))
        _apply(ws.cell(row=row, column=depot_col + 1, value=depot_reel))
        ecart_cell = ws.cell(row=row, column=depot_col + 2, value=ecart)
        ecart_fill = PAID_FILL if ecart >= 0 else LATE_FILL
        _apply(ecart_cell, fill=ecart_fill)

    # Freeze header rows
    ws.freeze_panes = "B3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"gestion_immobiliere_{year}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)
