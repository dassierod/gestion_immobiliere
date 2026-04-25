import os
import io
import re
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# App configuration
# ---------------------------------------------------------------------------
basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'instance', 'gestion_immo.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'uploads')

os.makedirs(os.path.join(basedir, 'instance'), exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

MOIS_NOMS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
             'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

PHONE_WARNING_MSG = 'Format de téléphone invalide. Utilisez le format +237XXXXXXXXX.'

db = SQLAlchemy(app)

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(30), default='proprietaire')  # gestionnaire | proprietaire
    nom = db.Column(db.String(100))
    prenom = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'{self.username} ({self.role})'


class Proprietaire(db.Model):
    __tablename__ = 'proprietaires'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(20))
    adresse = db.Column(db.String(255))
    biens = db.relationship('Bien', backref='proprietaire', lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'{self.prenom} {self.nom}'


class Bien(db.Model):
    __tablename__ = 'biens'
    id = db.Column(db.Integer, primary_key=True)
    reference = db.Column(db.String(50), unique=True, nullable=False)
    type_bien = db.Column(db.String(50), nullable=False)
    adresse = db.Column(db.String(255), nullable=False)
    ville = db.Column(db.String(100))
    code_postal = db.Column(db.String(10))
    surface = db.Column(db.Float)
    nb_pieces = db.Column(db.Integer)
    loyer_mensuel = db.Column(db.Float)
    charges = db.Column(db.Float, default=0)
    statut = db.Column(db.String(30), default='Disponible')
    description = db.Column(db.Text)
    proprietaire_id = db.Column(db.Integer, db.ForeignKey('proprietaires.id'))
    contrats = db.relationship('Contrat', backref='bien', lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def loyer_total(self):
        return (self.loyer_mensuel or 0) + (self.charges or 0)

    def __repr__(self):
        return f'{self.reference} - {self.adresse}'


class Locataire(db.Model):
    __tablename__ = 'locataires'
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(20))
    adresse = db.Column(db.String(255))
    date_naissance = db.Column(db.Date)
    profession = db.Column(db.String(100))
    revenu_mensuel = db.Column(db.Float)
    contrats = db.relationship('Contrat', backref='locataire', lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'{self.prenom} {self.nom}'


class Contrat(db.Model):
    __tablename__ = 'contrats'
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(50), unique=True, nullable=False)
    bien_id = db.Column(db.Integer, db.ForeignKey('biens.id'), nullable=False)
    locataire_id = db.Column(db.Integer, db.ForeignKey('locataires.id'), nullable=False)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date)
    loyer = db.Column(db.Float, nullable=False)
    charges = db.Column(db.Float, default=0)
    caution = db.Column(db.Float, default=0)
    statut = db.Column(db.String(30), default='Actif')
    paiements = db.relationship('Paiement', backref='contrat', lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def total_paye(self):
        return sum(p.montant for p in self.paiements if p.statut == 'Payé')

    def loyer_total(self):
        return (self.loyer or 0) + (self.charges or 0)

    def __repr__(self):
        return f'Contrat {self.numero}'


class Paiement(db.Model):
    __tablename__ = 'paiements'
    id = db.Column(db.Integer, primary_key=True)
    contrat_id = db.Column(db.Integer, db.ForeignKey('contrats.id'), nullable=False)
    date_paiement = db.Column(db.Date, nullable=False)
    montant = db.Column(db.Float, nullable=False)
    type_paiement = db.Column(db.String(50), default='Loyer')
    mode_paiement = db.Column(db.String(50), default='Virement')
    statut = db.Column(db.String(30), default='Payé')
    reference = db.Column(db.String(100))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'Paiement {self.id} - {self.montant} FCFA'


class Virement(db.Model):
    __tablename__ = 'virements'
    id = db.Column(db.Integer, primary_key=True)
    proprietaire_id = db.Column(db.Integer, db.ForeignKey('proprietaires.id'), nullable=False)
    montant = db.Column(db.Float, nullable=False)
    date_virement = db.Column(db.Date, nullable=False)
    mode_virement = db.Column(db.String(50), default='Mobile Money')
    reference = db.Column(db.String(100))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    proprietaire = db.relationship('Proprietaire', backref='virements')

    def __repr__(self):
        return f'Virement {self.id} - {self.montant} FCFA'


# ---------------------------------------------------------------------------
# Jinja2 filters
# ---------------------------------------------------------------------------

@app.template_filter('fcfa')
def format_fcfa(value):
    try:
        value = int(round(float(value)))
        formatted = f'{value:,}'.replace(',', '\u00a0')  # non-breaking space
        return f'{formatted} FCFA'
    except (ValueError, TypeError):
        return '0 FCFA'


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def validate_phone(phone):
    """Returns True if phone is empty or matches Cameroonian format."""
    if not phone:
        return True
    return bool(re.match(r'^\+237[0-9]{9}$', phone.replace(' ', '')))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Veuillez vous connecter pour accéder à cette page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def gestionnaire_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Veuillez vous connecter pour accéder à cette page.', 'warning')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or user.role != 'gestionnaire':
            flash('Accès réservé aux gestionnaires.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Context helpers
# ---------------------------------------------------------------------------

@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}


@app.context_processor
def inject_current_user():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        return {'current_user': user}
    return {'current_user': None}


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            flash(f'Bienvenue, {user.prenom or user.username} !', 'success')
            return redirect(url_for('dashboard'))
        flash('Identifiant ou mot de passe incorrect.', 'danger')
    return render_template('auth/login.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def dashboard():
    nb_biens = Bien.query.count()
    nb_biens_loues = Bien.query.filter_by(statut='Loué').count()
    nb_biens_disponibles = Bien.query.filter_by(statut='Disponible').count()
    nb_locataires = Locataire.query.count()
    nb_contrats_actifs = Contrat.query.filter_by(statut='Actif').count()

    today = date.today()
    debut_mois = today.replace(day=1)
    debut_annee = today.replace(month=1, day=1)

    revenus_mois = db.session.query(func.sum(Paiement.montant)).filter(
        Paiement.date_paiement >= debut_mois,
        Paiement.statut == 'Payé'
    ).scalar() or 0

    revenus_total = db.session.query(func.sum(Paiement.montant)).filter(
        Paiement.statut == 'Payé'
    ).scalar() or 0

    revenus_annee = db.session.query(func.sum(Paiement.montant)).filter(
        Paiement.date_paiement >= debut_annee,
        Paiement.statut == 'Payé'
    ).scalar() or 0

    montant_transfere_total = db.session.query(func.sum(Virement.montant)).scalar() or 0
    montant_transfere_annee = db.session.query(func.sum(Virement.montant)).filter(
        Virement.date_virement >= debut_annee
    ).scalar() or 0

    solde_a_transferer = revenus_total - montant_transfere_total

    paiements_retard = Paiement.query.filter_by(statut='Retard').all()
    paiements_retard_count = len(paiements_retard)
    paiements_retard_montant = sum(p.montant for p in paiements_retard)

    paiements_avance_count = Paiement.query.filter_by(statut='Avance').count()

    paiements_recents = Paiement.query.order_by(Paiement.created_at.desc()).limit(10).all()

    in_60_days = today + timedelta(days=60)
    contrats_expirant = Contrat.query.filter(
        Contrat.statut == 'Actif',
        Contrat.date_fin != None,
        Contrat.date_fin <= in_60_days,
        Contrat.date_fin >= today
    ).all()

    loyers_en_attente = Paiement.query.filter_by(statut='En attente').count()

    stats = {
        'nb_biens': nb_biens,
        'nb_biens_loues': nb_biens_loues,
        'nb_biens_disponibles': nb_biens_disponibles,
        'nb_locataires': nb_locataires,
        'nb_contrats_actifs': nb_contrats_actifs,
        'revenus_mois': revenus_mois,
        'revenus_total': revenus_total,
        'revenus_annee': revenus_annee,
        'montant_transfere_total': montant_transfere_total,
        'montant_transfere_annee': montant_transfere_annee,
        'solde_a_transferer': solde_a_transferer,
        'paiements_retard_count': paiements_retard_count,
        'paiements_retard_montant': paiements_retard_montant,
        'paiements_avance_count': paiements_avance_count,
        'loyers_en_attente': loyers_en_attente,
        'nb_contrats_expirant': len(contrats_expirant),
    }
    return render_template('dashboard.html', stats=stats,
                           paiements_recents=paiements_recents,
                           contrats_expirant=contrats_expirant)


# ---------------------------------------------------------------------------
# Biens (Properties)
# ---------------------------------------------------------------------------

@app.route('/biens')
@login_required
def biens_list():
    statut = request.args.get('statut', '')
    search = request.args.get('q', '')
    query = Bien.query
    if statut:
        query = query.filter_by(statut=statut)
    if search:
        query = query.filter(
            db.or_(Bien.reference.ilike(f'%{search}%'),
                   Bien.adresse.ilike(f'%{search}%'),
                   Bien.ville.ilike(f'%{search}%'))
        )
    biens = query.order_by(Bien.created_at.desc()).all()
    return render_template('biens/list.html', biens=biens, statut=statut, search=search)


@app.route('/biens/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def bien_create():
    proprietaires = Proprietaire.query.order_by(Proprietaire.nom).all()
    if request.method == 'POST':
        bien = Bien(
            reference=request.form['reference'],
            type_bien=request.form['type_bien'],
            adresse=request.form['adresse'],
            ville=request.form.get('ville', ''),
            code_postal=request.form.get('code_postal', ''),
            surface=float(request.form['surface']) if request.form.get('surface') else None,
            nb_pieces=int(request.form['nb_pieces']) if request.form.get('nb_pieces') else None,
            loyer_mensuel=float(request.form['loyer_mensuel']) if request.form.get('loyer_mensuel') else None,
            charges=float(request.form['charges']) if request.form.get('charges') else 0,
            statut=request.form.get('statut', 'Disponible'),
            description=request.form.get('description', ''),
            proprietaire_id=int(request.form['proprietaire_id']) if request.form.get('proprietaire_id') else None,
        )
        db.session.add(bien)
        db.session.commit()
        flash('Bien immobilier créé avec succès.', 'success')
        return redirect(url_for('biens_list'))
    return render_template('biens/form.html', bien=None, proprietaires=proprietaires)


@app.route('/biens/<int:bien_id>')
@login_required
def bien_detail(bien_id):
    bien = Bien.query.get_or_404(bien_id)
    contrats = Contrat.query.filter_by(bien_id=bien_id).order_by(Contrat.date_debut.desc()).all()
    return render_template('biens/detail.html', bien=bien, contrats=contrats)


@app.route('/biens/<int:bien_id>/modifier', methods=['GET', 'POST'])
@gestionnaire_required
def bien_edit(bien_id):
    bien = Bien.query.get_or_404(bien_id)
    proprietaires = Proprietaire.query.order_by(Proprietaire.nom).all()
    if request.method == 'POST':
        bien.reference = request.form['reference']
        bien.type_bien = request.form['type_bien']
        bien.adresse = request.form['adresse']
        bien.ville = request.form.get('ville', '')
        bien.code_postal = request.form.get('code_postal', '')
        bien.surface = float(request.form['surface']) if request.form.get('surface') else None
        bien.nb_pieces = int(request.form['nb_pieces']) if request.form.get('nb_pieces') else None
        bien.loyer_mensuel = float(request.form['loyer_mensuel']) if request.form.get('loyer_mensuel') else None
        bien.charges = float(request.form['charges']) if request.form.get('charges') else 0
        bien.statut = request.form.get('statut', 'Disponible')
        bien.description = request.form.get('description', '')
        bien.proprietaire_id = int(request.form['proprietaire_id']) if request.form.get('proprietaire_id') else None
        db.session.commit()
        flash('Bien immobilier mis à jour.', 'success')
        return redirect(url_for('bien_detail', bien_id=bien.id))
    return render_template('biens/form.html', bien=bien, proprietaires=proprietaires)


@app.route('/biens/<int:bien_id>/supprimer', methods=['POST'])
@gestionnaire_required
def bien_delete(bien_id):
    bien = Bien.query.get_or_404(bien_id)
    db.session.delete(bien)
    db.session.commit()
    flash('Bien immobilier supprimé.', 'warning')
    return redirect(url_for('biens_list'))


# ---------------------------------------------------------------------------
# Locataires (Tenants)
# ---------------------------------------------------------------------------

@app.route('/locataires')
@login_required
def locataires_list():
    search = request.args.get('q', '')
    query = Locataire.query
    if search:
        query = query.filter(
            db.or_(Locataire.nom.ilike(f'%{search}%'),
                   Locataire.prenom.ilike(f'%{search}%'),
                   Locataire.email.ilike(f'%{search}%'))
        )
    locataires = query.order_by(Locataire.nom).all()
    return render_template('locataires/list.html', locataires=locataires, search=search)


@app.route('/locataires/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def locataire_create():
    if request.method == 'POST':
        dob = None
        if request.form.get('date_naissance'):
            dob = datetime.strptime(request.form['date_naissance'], '%Y-%m-%d').date()
        phone = request.form.get('telephone', '')
        if not validate_phone(phone):
            flash(PHONE_WARNING_MSG, 'warning')
        locataire = Locataire(
            nom=request.form['nom'],
            prenom=request.form['prenom'],
            email=request.form.get('email', ''),
            telephone=request.form.get('telephone', ''),
            adresse=request.form.get('adresse', ''),
            date_naissance=dob,
            profession=request.form.get('profession', ''),
            revenu_mensuel=float(request.form['revenu_mensuel']) if request.form.get('revenu_mensuel') else None,
        )
        db.session.add(locataire)
        db.session.commit()
        flash('Locataire créé avec succès.', 'success')
        return redirect(url_for('locataires_list'))
    return render_template('locataires/form.html', locataire=None)


@app.route('/locataires/<int:loc_id>')
@login_required
def locataire_detail(loc_id):
    locataire = Locataire.query.get_or_404(loc_id)
    contrats = Contrat.query.filter_by(locataire_id=loc_id).order_by(Contrat.date_debut.desc()).all()
    return render_template('locataires/detail.html', locataire=locataire, contrats=contrats)


@app.route('/locataires/<int:loc_id>/modifier', methods=['GET', 'POST'])
@gestionnaire_required
def locataire_edit(loc_id):
    locataire = Locataire.query.get_or_404(loc_id)
    if request.method == 'POST':
        dob = None
        if request.form.get('date_naissance'):
            dob = datetime.strptime(request.form['date_naissance'], '%Y-%m-%d').date()
        phone = request.form.get('telephone', '')
        if not validate_phone(phone):
            flash(PHONE_WARNING_MSG, 'warning')
        locataire.nom = request.form['nom']
        locataire.prenom = request.form['prenom']
        locataire.email = request.form.get('email', '')
        locataire.telephone = request.form.get('telephone', '')
        locataire.adresse = request.form.get('adresse', '')
        locataire.date_naissance = dob
        locataire.profession = request.form.get('profession', '')
        locataire.revenu_mensuel = float(request.form['revenu_mensuel']) if request.form.get('revenu_mensuel') else None
        db.session.commit()
        flash('Locataire mis à jour.', 'success')
        return redirect(url_for('locataire_detail', loc_id=locataire.id))
    return render_template('locataires/form.html', locataire=locataire)


@app.route('/locataires/<int:loc_id>/supprimer', methods=['POST'])
@gestionnaire_required
def locataire_delete(loc_id):
    locataire = Locataire.query.get_or_404(loc_id)
    db.session.delete(locataire)
    db.session.commit()
    flash('Locataire supprimé.', 'warning')
    return redirect(url_for('locataires_list'))


# ---------------------------------------------------------------------------
# Contrats (Contracts)
# ---------------------------------------------------------------------------

@app.route('/contrats')
@login_required
def contrats_list():
    statut = request.args.get('statut', '')
    query = Contrat.query
    if statut:
        query = query.filter_by(statut=statut)
    contrats = query.order_by(Contrat.date_debut.desc()).all()
    return render_template('contrats/list.html', contrats=contrats, statut=statut)


@app.route('/contrats/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def contrat_create():
    biens = Bien.query.filter(Bien.statut.in_(['Disponible', 'Loué'])).order_by(Bien.reference).all()
    locataires = Locataire.query.order_by(Locataire.nom).all()
    if request.method == 'POST':
        date_fin = None
        if request.form.get('date_fin'):
            date_fin = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date()
        contrat = Contrat(
            numero=request.form['numero'],
            bien_id=int(request.form['bien_id']),
            locataire_id=int(request.form['locataire_id']),
            date_debut=datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date(),
            date_fin=date_fin,
            loyer=float(request.form['loyer']),
            charges=float(request.form['charges']) if request.form.get('charges') else 0,
            caution=float(request.form['caution']) if request.form.get('caution') else 0,
            statut=request.form.get('statut', 'Actif'),
        )
        db.session.add(contrat)
        # Update bien status
        bien = Bien.query.get(contrat.bien_id)
        if bien and contrat.statut == 'Actif':
            bien.statut = 'Loué'
        db.session.commit()
        flash('Contrat créé avec succès.', 'success')
        return redirect(url_for('contrats_list'))
    return render_template('contrats/form.html', contrat=None, biens=biens, locataires=locataires)


@app.route('/contrats/<int:contrat_id>')
@login_required
def contrat_detail(contrat_id):
    contrat = Contrat.query.get_or_404(contrat_id)
    paiements = Paiement.query.filter_by(contrat_id=contrat_id).order_by(Paiement.date_paiement.desc()).all()
    return render_template('contrats/detail.html', contrat=contrat, paiements=paiements)


@app.route('/contrats/<int:contrat_id>/modifier', methods=['GET', 'POST'])
@gestionnaire_required
def contrat_edit(contrat_id):
    contrat = Contrat.query.get_or_404(contrat_id)
    biens = Bien.query.order_by(Bien.reference).all()
    locataires = Locataire.query.order_by(Locataire.nom).all()
    if request.method == 'POST':
        date_fin = None
        if request.form.get('date_fin'):
            date_fin = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date()
        contrat.numero = request.form['numero']
        contrat.bien_id = int(request.form['bien_id'])
        contrat.locataire_id = int(request.form['locataire_id'])
        contrat.date_debut = datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date()
        contrat.date_fin = date_fin
        contrat.loyer = float(request.form['loyer'])
        contrat.charges = float(request.form['charges']) if request.form.get('charges') else 0
        contrat.caution = float(request.form['caution']) if request.form.get('caution') else 0
        old_statut = contrat.statut
        contrat.statut = request.form.get('statut', 'Actif')
        # Update bien status if contract status changed
        bien = Bien.query.get(contrat.bien_id)
        if bien:
            if contrat.statut == 'Actif':
                bien.statut = 'Loué'
            elif old_statut == 'Actif' and contrat.statut != 'Actif':
                bien.statut = 'Disponible'
        db.session.commit()
        flash('Contrat mis à jour.', 'success')
        return redirect(url_for('contrat_detail', contrat_id=contrat.id))
    return render_template('contrats/form.html', contrat=contrat, biens=biens, locataires=locataires)


@app.route('/contrats/<int:contrat_id>/supprimer', methods=['POST'])
@gestionnaire_required
def contrat_delete(contrat_id):
    contrat = Contrat.query.get_or_404(contrat_id)
    db.session.delete(contrat)
    db.session.commit()
    flash('Contrat supprimé.', 'warning')
    return redirect(url_for('contrats_list'))


# ---------------------------------------------------------------------------
# Paiements (Payments)
# ---------------------------------------------------------------------------

@app.route('/paiements')
@login_required
def paiements_list():
    statut = request.args.get('statut', '')
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')
    locataire_id = request.args.get('locataire_id', '')
    bien_id = request.args.get('bien_id', '')

    query = Paiement.query.join(Contrat)
    if statut:
        query = query.filter(Paiement.statut == statut)
    if date_debut:
        query = query.filter(Paiement.date_paiement >= datetime.strptime(date_debut, '%Y-%m-%d').date())
    if date_fin:
        query = query.filter(Paiement.date_paiement <= datetime.strptime(date_fin, '%Y-%m-%d').date())
    if locataire_id:
        query = query.filter(Contrat.locataire_id == int(locataire_id))
    if bien_id:
        query = query.filter(Contrat.bien_id == int(bien_id))

    paiements = query.order_by(Paiement.date_paiement.desc()).all()
    locataires = Locataire.query.order_by(Locataire.nom).all()
    biens = Bien.query.order_by(Bien.reference).all()
    return render_template('paiements/list.html', paiements=paiements, statut=statut,
                           date_debut=date_debut, date_fin=date_fin,
                           locataire_id=locataire_id, bien_id=bien_id,
                           locataires=locataires, biens=biens)


@app.route('/paiements/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def paiement_create():
    contrats = Contrat.query.filter_by(statut='Actif').order_by(Contrat.numero).all()
    if request.method == 'POST':
        paiement = Paiement(
            contrat_id=int(request.form['contrat_id']),
            date_paiement=datetime.strptime(request.form['date_paiement'], '%Y-%m-%d').date(),
            montant=float(request.form['montant']),
            type_paiement=request.form.get('type_paiement', 'Loyer'),
            mode_paiement=request.form.get('mode_paiement', 'Mobile Money'),
            statut=request.form.get('statut', 'Payé'),
            reference=request.form.get('reference', ''),
            note=request.form.get('note', ''),
        )
        db.session.add(paiement)
        db.session.commit()
        flash('Paiement enregistré avec succès.', 'success')
        return redirect(url_for('paiements_list'))
    return render_template('paiements/form.html', paiement=None, contrats=contrats)


@app.route('/paiements/<int:paiement_id>/modifier', methods=['GET', 'POST'])
@gestionnaire_required
def paiement_edit(paiement_id):
    paiement = Paiement.query.get_or_404(paiement_id)
    contrats = Contrat.query.order_by(Contrat.numero).all()
    if request.method == 'POST':
        paiement.contrat_id = int(request.form['contrat_id'])
        paiement.date_paiement = datetime.strptime(request.form['date_paiement'], '%Y-%m-%d').date()
        paiement.montant = float(request.form['montant'])
        paiement.type_paiement = request.form.get('type_paiement', 'Loyer')
        paiement.mode_paiement = request.form.get('mode_paiement', 'Mobile Money')
        paiement.statut = request.form.get('statut', 'Payé')
        paiement.reference = request.form.get('reference', '')
        paiement.note = request.form.get('note', '')
        db.session.commit()
        flash('Paiement mis à jour.', 'success')
        return redirect(url_for('paiements_list'))
    return render_template('paiements/form.html', paiement=paiement, contrats=contrats)


@app.route('/paiements/<int:paiement_id>/supprimer', methods=['POST'])
@gestionnaire_required
def paiement_delete(paiement_id):
    paiement = Paiement.query.get_or_404(paiement_id)
    db.session.delete(paiement)
    db.session.commit()
    flash('Paiement supprimé.', 'warning')
    return redirect(url_for('paiements_list'))


# ---------------------------------------------------------------------------
# Proprietaires (Owners)
# ---------------------------------------------------------------------------

@app.route('/proprietaires')
@login_required
def proprietaires_list():
    proprietaires = Proprietaire.query.order_by(Proprietaire.nom).all()
    return render_template('proprietaires/list.html', proprietaires=proprietaires)


@app.route('/proprietaires/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def proprietaire_create():
    if request.method == 'POST':
        phone = request.form.get('telephone', '')
        if not validate_phone(phone):
            flash(PHONE_WARNING_MSG, 'warning')
        proprietaire = Proprietaire(
            nom=request.form['nom'],
            prenom=request.form['prenom'],
            email=request.form.get('email', ''),
            telephone=request.form.get('telephone', ''),
            adresse=request.form.get('adresse', ''),
        )
        db.session.add(proprietaire)
        db.session.commit()
        flash('Propriétaire créé avec succès.', 'success')
        return redirect(url_for('proprietaires_list'))
    return render_template('proprietaires/form.html', proprietaire=None)


@app.route('/proprietaires/<int:prop_id>/modifier', methods=['GET', 'POST'])
@gestionnaire_required
def proprietaire_edit(prop_id):
    proprietaire = Proprietaire.query.get_or_404(prop_id)
    if request.method == 'POST':
        phone = request.form.get('telephone', '')
        if not validate_phone(phone):
            flash(PHONE_WARNING_MSG, 'warning')
        proprietaire.nom = request.form['nom']
        proprietaire.prenom = request.form['prenom']
        proprietaire.email = request.form.get('email', '')
        proprietaire.telephone = request.form.get('telephone', '')
        proprietaire.adresse = request.form.get('adresse', '')
        db.session.commit()
        flash('Propriétaire mis à jour.', 'success')
        return redirect(url_for('proprietaires_list'))
    return render_template('proprietaires/form.html', proprietaire=proprietaire)


@app.route('/proprietaires/<int:prop_id>/supprimer', methods=['POST'])
@gestionnaire_required
def proprietaire_delete(prop_id):
    proprietaire = Proprietaire.query.get_or_404(prop_id)
    db.session.delete(proprietaire)
    db.session.commit()
    flash('Propriétaire supprimé.', 'warning')
    return redirect(url_for('proprietaires_list'))


# ---------------------------------------------------------------------------
# Excel Import / Export
# ---------------------------------------------------------------------------

def _style_header_row(ws, row, col_count):
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _add_borders(ws, min_row, max_row, max_col):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = border


@app.route('/export/excel')
@login_required
def export_excel():
    wb = openpyxl.Workbook()

    # --- Biens ---
    ws_biens = wb.active
    ws_biens.title = 'Biens'
    headers_biens = ['Référence', 'Type', 'Adresse', 'Ville', 'Code Postal',
                     'Surface (m²)', 'Nb Pièces', 'Loyer (FCFA)', 'Charges (FCFA)', 'Statut', 'Description']
    ws_biens.append(headers_biens)
    _style_header_row(ws_biens, 1, len(headers_biens))
    for bien in Bien.query.order_by(Bien.reference).all():
        ws_biens.append([bien.reference, bien.type_bien, bien.adresse, bien.ville,
                         bien.code_postal, bien.surface, bien.nb_pieces,
                         bien.loyer_mensuel, bien.charges, bien.statut, bien.description])
    for col in range(1, len(headers_biens) + 1):
        ws_biens.column_dimensions[get_column_letter(col)].width = 18
    _add_borders(ws_biens, 1, ws_biens.max_row, len(headers_biens))

    # --- Locataires ---
    ws_loc = wb.create_sheet('Locataires')
    headers_loc = ['Nom', 'Prénom', 'Email', 'Téléphone', 'Adresse',
                   'Date de naissance', 'Profession', 'Revenu mensuel (FCFA)']
    ws_loc.append(headers_loc)
    _style_header_row(ws_loc, 1, len(headers_loc))
    for loc in Locataire.query.order_by(Locataire.nom).all():
        ws_loc.append([loc.nom, loc.prenom, loc.email, loc.telephone, loc.adresse,
                       loc.date_naissance.strftime('%d/%m/%Y') if loc.date_naissance else '',
                       loc.profession, loc.revenu_mensuel])
    for col in range(1, len(headers_loc) + 1):
        ws_loc.column_dimensions[get_column_letter(col)].width = 20
    _add_borders(ws_loc, 1, ws_loc.max_row, len(headers_loc))

    # --- Contrats ---
    ws_cont = wb.create_sheet('Contrats')
    headers_cont = ['Numéro', 'Bien (Réf)', 'Locataire', 'Date début', 'Date fin',
                    'Loyer (FCFA)', 'Charges (FCFA)', 'Caution (FCFA)', 'Statut']
    ws_cont.append(headers_cont)
    _style_header_row(ws_cont, 1, len(headers_cont))
    for c in Contrat.query.order_by(Contrat.date_debut.desc()).all():
        ws_cont.append([
            c.numero, c.bien.reference if c.bien else '',
            str(c.locataire) if c.locataire else '',
            c.date_debut.strftime('%d/%m/%Y') if c.date_debut else '',
            c.date_fin.strftime('%d/%m/%Y') if c.date_fin else '',
            c.loyer, c.charges, c.caution, c.statut
        ])
    for col in range(1, len(headers_cont) + 1):
        ws_cont.column_dimensions[get_column_letter(col)].width = 20
    _add_borders(ws_cont, 1, ws_cont.max_row, len(headers_cont))

    # --- Paiements ---
    ws_pay = wb.create_sheet('Paiements')
    headers_pay = ['Contrat', 'Date', 'Montant (FCFA)', 'Type', 'Mode', 'Statut', 'Référence', 'Note']
    ws_pay.append(headers_pay)
    _style_header_row(ws_pay, 1, len(headers_pay))
    for p in Paiement.query.order_by(Paiement.date_paiement.desc()).all():
        ws_pay.append([
            p.contrat.numero if p.contrat else '',
            p.date_paiement.strftime('%d/%m/%Y') if p.date_paiement else '',
            p.montant, p.type_paiement, p.mode_paiement, p.statut, p.reference, p.note
        ])
    for col in range(1, len(headers_pay) + 1):
        ws_pay.column_dimensions[get_column_letter(col)].width = 18
    _add_borders(ws_pay, 1, ws_pay.max_row, len(headers_pay))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'gestion_immobiliere_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/import/excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(url_for('import_excel'))
        f = request.files['file']
        if f.filename == '':
            flash('Aucun fichier sélectionné.', 'danger')
            return redirect(url_for('import_excel'))
        if not f.filename.endswith(('.xlsx', '.xls')):
            flash('Seuls les fichiers Excel (.xlsx, .xls) sont acceptés.', 'danger')
            return redirect(url_for('import_excel'))

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            imported = {'biens': 0, 'locataires': 0, 'contrats': 0, 'paiements': 0}

            # Import Biens
            if 'Biens' in wb.sheetnames:
                ws = wb['Biens']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    existing = Bien.query.filter_by(reference=str(row[0])).first()
                    if not existing:
                        bien = Bien(
                            reference=str(row[0]),
                            type_bien=str(row[1]) if row[1] else 'Appartement',
                            adresse=str(row[2]) if row[2] else '',
                            ville=str(row[3]) if row[3] else '',
                            code_postal=str(row[4]) if row[4] else '',
                            surface=float(row[5]) if row[5] else None,
                            nb_pieces=int(row[6]) if row[6] else None,
                            loyer_mensuel=float(row[7]) if row[7] else None,
                            charges=float(row[8]) if row[8] else 0,
                            statut=str(row[9]) if row[9] else 'Disponible',
                            description=str(row[10]) if row[10] else '',
                        )
                        db.session.add(bien)
                        imported['biens'] += 1

            # Import Locataires
            if 'Locataires' in wb.sheetnames:
                ws = wb['Locataires']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    dob = None
                    if row[5]:
                        try:
                            if isinstance(row[5], str):
                                dob = datetime.strptime(row[5], '%d/%m/%Y').date()
                            elif hasattr(row[5], 'date'):
                                dob = row[5].date()
                        except (ValueError, AttributeError):
                            pass
                    locataire = Locataire(
                        nom=str(row[0]),
                        prenom=str(row[1]) if row[1] else '',
                        email=str(row[2]) if row[2] else '',
                        telephone=str(row[3]) if row[3] else '',
                        adresse=str(row[4]) if row[4] else '',
                        date_naissance=dob,
                        profession=str(row[6]) if row[6] else '',
                        revenu_mensuel=float(row[7]) if row[7] else None,
                    )
                    db.session.add(locataire)
                    imported['locataires'] += 1

            db.session.commit()
            flash(f"Import réussi : {imported['biens']} bien(s), {imported['locataires']} locataire(s).", 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'import : {str(e)}', 'danger')

        return redirect(url_for('dashboard'))
    return render_template('import.html')


# ---------------------------------------------------------------------------
# Seed demo data
# ---------------------------------------------------------------------------

@app.route('/seed')
def seed_demo():
    if Bien.query.count() > 0:
        flash('Des données existent déjà.', 'info')
        return redirect(url_for('dashboard'))

    # Create default admin user
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', role='gestionnaire', nom='Administrateur', prenom='Système')
        admin.set_password('admin123')
        db.session.add(admin)

    # Proprietaire
    p1 = Proprietaire(nom='Mbarga', prenom='Jean', email='jean.mbarga@email.cm',
                      telephone='+237677123456', adresse='Bastos, Yaoundé, Cameroun')
    db.session.add(p1)
    db.session.flush()

    # Create a proprietaire user
    if not User.query.filter_by(username='mbarga').first():
        prop_user = User(username='mbarga', role='proprietaire', nom='Mbarga', prenom='Jean')
        prop_user.set_password('mbarga123')
        db.session.add(prop_user)

    # Biens
    b1 = Bien(reference='APP-001', type_bien='Appartement', adresse='Rue Nachtigal, Quartier du Lac',
              ville='Yaoundé', code_postal='', surface=65.0, nb_pieces=3,
              loyer_mensuel=150000.0, charges=15000.0, statut='Loué', proprietaire_id=p1.id)
    b2 = Bien(reference='APP-002', type_bien='Appartement', adresse='Akwa, Boulevard de la Liberté',
              ville='Douala', code_postal='', surface=45.0, nb_pieces=2,
              loyer_mensuel=100000.0, charges=10000.0, statut='Disponible', proprietaire_id=p1.id)
    b3 = Bien(reference='MAI-001', type_bien='Maison', adresse='Omnisports, Rue 1.234',
              ville='Yaoundé', code_postal='', surface=120.0, nb_pieces=5,
              loyer_mensuel=300000.0, charges=25000.0, statut='Loué', proprietaire_id=p1.id)
    db.session.add_all([b1, b2, b3])
    db.session.flush()

    # Locataires
    l1 = Locataire(nom='Nkomo', prenom='Marie', email='marie.nkomo@email.cm',
                   telephone='+237698765432', profession='Ingénieure', revenu_mensuel=450000.0)
    l2 = Locataire(nom='Fotso', prenom='Paul', email='paul.fotso@email.cm',
                   telephone='+237677112233', profession='Médecin', revenu_mensuel=700000.0)
    db.session.add_all([l1, l2])
    db.session.flush()

    # Contrats
    c1 = Contrat(numero='CTR-2024-001', bien_id=b1.id, locataire_id=l1.id,
                 date_debut=date(2024, 1, 1), date_fin=date(2026, 12, 31),
                 loyer=150000.0, charges=15000.0, caution=300000.0, statut='Actif')
    c2 = Contrat(numero='CTR-2024-002', bien_id=b3.id, locataire_id=l2.id,
                 date_debut=date(2024, 3, 1), date_fin=date(2027, 2, 28),
                 loyer=300000.0, charges=25000.0, caution=600000.0, statut='Actif')
    db.session.add_all([c1, c2])
    db.session.flush()

    # Paiements
    today = date.today()
    for i in range(1, 5):
        month = today.month - i
        year = today.year
        if month <= 0:
            month += 12
            year -= 1
        db.session.add(Paiement(contrat_id=c1.id, date_paiement=date(year, month, 5),
                                montant=165000.0, type_paiement='Loyer',
                                mode_paiement='Mobile Money', statut='Payé'))
        db.session.add(Paiement(contrat_id=c2.id, date_paiement=date(year, month, 3),
                                montant=325000.0, type_paiement='Loyer',
                                mode_paiement='Virement bancaire', statut='Payé'))
    db.session.add(Paiement(contrat_id=c1.id, date_paiement=date(today.year, today.month, 5),
                            montant=165000.0, type_paiement='Loyer',
                            mode_paiement='Mobile Money', statut='En attente'))

    # Virement demo
    db.session.add(Virement(proprietaire_id=p1.id, montant=980000.0,
                            date_virement=date(today.year - 1 if today.month == 1 else today.year,
                                              today.month - 1 if today.month > 1 else 12, 15),
                            mode_virement='Virement bancaire', reference='VIR-2024-001',
                            note='Virement trimestriel'))

    db.session.commit()
    flash('Données de démonstration créées avec succès !', 'success')
    return redirect(url_for('dashboard'))


# ---------------------------------------------------------------------------
# Virements (Owner transfers)
# ---------------------------------------------------------------------------

@app.route('/virements')
@login_required
def virements_list():
    virements = Virement.query.order_by(Virement.date_virement.desc()).all()
    total = sum(v.montant for v in virements)
    return render_template('virements/list.html', virements=virements, total=total)


@app.route('/virements/nouveau', methods=['GET', 'POST'])
@gestionnaire_required
def virement_create():
    proprietaires = Proprietaire.query.order_by(Proprietaire.nom).all()
    if request.method == 'POST':
        virement = Virement(
            proprietaire_id=int(request.form['proprietaire_id']),
            montant=float(request.form['montant']),
            date_virement=datetime.strptime(request.form['date_virement'], '%Y-%m-%d').date(),
            mode_virement=request.form.get('mode_virement', 'Mobile Money'),
            reference=request.form.get('reference', ''),
            note=request.form.get('note', ''),
        )
        db.session.add(virement)
        db.session.commit()
        flash('Virement enregistré avec succès.', 'success')
        return redirect(url_for('virements_list'))
    return render_template('virements/form.html', virement=None, proprietaires=proprietaires)


@app.route('/virements/<int:virement_id>/supprimer', methods=['POST'])
@gestionnaire_required
def virement_delete(virement_id):
    virement = Virement.query.get_or_404(virement_id)
    db.session.delete(virement)
    db.session.commit()
    flash('Virement supprimé.', 'warning')
    return redirect(url_for('virements_list'))


# ---------------------------------------------------------------------------
# Rapports (Reports)
# ---------------------------------------------------------------------------


@app.route('/rapports')
@login_required
def rapports():
    annee_courante = date.today().year
    annee = int(request.args.get('annee', annee_courante))
    annees_disponibles = list(range(annee_courante, annee_courante - 5, -1))

    rapports_data = []
    total_revenus = 0
    total_virements = 0

    for mois in range(1, 13):
        debut = date(annee, mois, 1)
        if mois == 12:
            fin = date(annee, 12, 31)
        else:
            fin = date(annee, mois + 1, 1) - timedelta(days=1)

        revenus = db.session.query(func.sum(Paiement.montant)).filter(
            Paiement.statut == 'Payé',
            Paiement.date_paiement >= debut,
            Paiement.date_paiement <= fin
        ).scalar() or 0

        virements = db.session.query(func.sum(Virement.montant)).filter(
            Virement.date_virement >= debut,
            Virement.date_virement <= fin
        ).scalar() or 0

        rapports_data.append({
            'mois': MOIS_NOMS[mois - 1],
            'revenus': revenus,
            'virements': virements,
            'solde': revenus - virements,
        })
        total_revenus += revenus
        total_virements += virements

    return render_template('rapports/index.html',
                           rapports_data=rapports_data,
                           annee=annee,
                           annees_disponibles=annees_disponibles,
                           total_revenus=total_revenus,
                           total_virements=total_virements,
                           total_solde=total_revenus - total_virements)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

with app.app_context():
    db.create_all()
    # Create default admin if no users exist
    if not User.query.first():
        admin = User(username='admin', role='gestionnaire', nom='Administrateur', prenom='Système')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()

if __name__ == '__main__':
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=5000)
